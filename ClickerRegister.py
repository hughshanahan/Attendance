from zipfile import ZipFile as _ZipFile 
from xml.dom import minidom as _minidom
from pathlib import Path
import openpyxl
import re
import copy
import os
import csv
import shutil

"""
Read in TurningPoint file and return list of clicker ID's

input: TurningPoint file name
output: list of clicker ID's
"""
def ClickerRegister(turningPointFileName):
    
    # extract TTSession.xml
    tpFile = _ZipFile(turningPointFileName)
    tpFile.extract("TTSession.xml")

    # read TTSession.xml file
    ttFile = open("TTSession.xml")
    
    xmldoc = _minidom.parseString(ttFile.readline())

    # get click device IDS 
    devices = [] 
    iParticipants = 0
    for child in xmldoc.getElementsByTagName("participant") : 
        # print child.childNodes[0].localName,child.childNodes[0].childNodes[0].nodeValue

        deviceID = child.childNodes[0].childNodes[0].nodeValue        
        devices.append(deviceID)

        iParticipants += 1 
        
#    print "Found ", iParticipants
    
    return(devices)
     

"""
 Return string with date of file name string
 Assume filename is of form dd-mm-yy hh-mm.tpxz
 input: filename
 output: dd-mm-yyyy
"""
def getDate(fName):
    bName = os.path.basename(fName)
    match = re.match(r'^\d\d-\d\d-\d\d\d\d',bName)
    if match:
        return bName[:10]
    else:
        raise Exception("getDate: filename should start with dd-mm-yyyy")
        
"""
 Return string of start time of lecture from a file name string
 Assume filename is of form dd-mm-yyyy hh-mm.tpxz
 input: filename
 output: hh
 
"""        
def getStartHour(fName):
    bName = os.path.basename(fName)
    match = re.match(r'^\d\d-\d\d-\d\d\d\d\s\d\d-\d\d.tpzx$',bName)
    if match:
        return bName[11:13]
    else:
        raise Exception("getStartHour: filename should have format dd-mm-yyyy hh-mm.tpzx")

"""
 getActivityRoot create end of string for activity
 input tpFn
 output string of form dd.<Mon>.yyyy_hh00
"""         

def getActivityRoot(turningPointFileName):
    hour = getStartHour(turningPointFileName)
    thisDateString = getDate(turningPointFileName)
    theseMonths = {"01":"Jan",
                    "02":"Feb",
                    "03":"Mar",
                    "04":"Apr",
                    "05":"May",
                    "06":"Jun",
                    "07":"Jul",
                    "08":"Aug",
                    "09":"Sep",
                    "10":"Oct",
                    "11":"Nov",
                    "12":"Dec"
                    }

    day = thisDateString[:2]
    month = theseMonths[thisDateString[3:5]]
    year = thisDateString[6:10]   
   
    activityRoot = day+"."+month+"."+year+"_"+hour+"00"
    return(activityRoot)



"""
 Return dict with student data stored in a spreadsheet
 in particular 
     keys are student ID's
     each value is a dict with keys of First Name, Surname, Courses (array)
 input: filename of spreadsheet
 output: dict as describe above    

"""        
def buildStudentDict(filename): 
    
# first worksheet has the useful data
    studentFile = openpyxl.load_workbook(filename,data_only=True)    
    ws = studentFile.worksheets[0]
    
    studentData = {}
    for row in ws.iter_rows(max_col=7,values_only=True):
        key = row[2]
        if len(row) > 0:
            if not key in studentData.keys() :
                studentData[key] = {}
                studentData[key]['First'] = row[4]
                studentData[key]['Surname'] = row[6]
                studentData[key]['clicker'] = row[3]
                studentData[key]['courses'] = [str(row[0]) + str(row[1]),]
            else:
                studentData[key]['courses'].append(str(row[0]) + str(row[1]))     
            
    return(studentData)

"""
Return subset of studentData where a student is attending a specific course

input: studentData (dict), course (string of from CS|IYNNNN)
output: list with keys of tudentData that has courses matching the course
"""

def selectCourseStudentData(data,course):
   
    attendingCourse = []
    for k,v in data.items():
        if course in v['courses']:
            attendingCourse.append(k)
            
    return(attendingCourse)    

"""
Return subset of studentData with a list of specific clicker ID's

input: studentData (dict), IDs (list)
output: list with keys of studentData that has clicker ID's matching the list of IDs
if there are any malformed clicker IDs, add the student ID to 
a separate list of student IDs 

"""

def selectIDStudentData(data,IDs):
   
    studentIDs = []
    malformedIDs = []
    for k in data.keys(): 
        if data[k] and type(data[k]) == dict:
            v = data[k]
            if v['clicker'] and type(v['clicker']) == str:
                if ( len(v['clicker']) == 6 ):
                    if  v['clicker'] in IDs:
                        studentIDs.append(k)
                else:
                    malformedIDs.append(k)
            else:
                malformedIDs.append(k)
        else:
            malformedIDs.append(k) 
                    
    return(studentIDs,malformedIDs)     
    
"""
Create a dict with studentData based on two lists; 1 list being students registered on a course; another
being students who attended the lecture
New dict has two news values for each key 
1 key is Present (True for present or has a malformed ID, False for absent)
1 key registered on course

"""    

def collateStudentsInLecture(allStudentData,registered,attended,malformed):
    r = set(registered)
    a = set(attended)
       
    rUa = r.union(a)
    thisStudentData = {}
    
    for i in rUa:
        thisStudentData[i] = copy.deepcopy(allStudentData[i])
        thisStudentData[i]['Present'] = i in attended or i in malformed
        thisStudentData[i]['Registered'] = i in registered
    
    
    return(thisStudentData)


"""
Create a spreadsheet with present/absent data


"""

def createAttendanceSpreadsheet(collated,date,course,activity,CRN,folder):
    wb = openpyxl.Workbook()
#    ws = wb.active()
    ws = wb.create_sheet("Attendance")
    names=("Activity",activity,"CRN",CRN)
    ws.append(names)
    names=["First name","Surname","Student ID","Clicker ID","Present",]
    ws.append(names)
    for k,v in collated.items():

        if v['Registered']:
            if v["Present"]:
                p = "Yes"
            else:
                p = "No"
            thisRow = [v['First'], v['Surname'], k, v['clicker'],p]
            ws.append(thisRow)
#    thisRow = ["Students not registered on this module but attended.",]
#    ws.append(thisRow)
        
#    for k,v in collated.items():
#        if v["Present"]:
#            if not v['Registered']:
#                p = "Yes"
#                thisRow = (v['First'], v['Surname'], k, v['clicker'],p)
#                ws.append(thisRow)               
     
    fname = course + "_" + date +".xlsx"   
            
    wb.save(Path(folder) / fname)
    
"""
Update a spreadsheet of the register report with present/absent data

"""

def updateRegisterReport(collated,activity,activityFn):
# Create new worksheet to store data
    tFn = "/Users/upac004/Downloads/tmp.csv"
    with open(tFn,'w',newline='') as csvOut:
        wsO = csv.writer(csvOut,delimiter=',',quotechar='"')
    
# first worksheet has the useful data
        with open(activityFn, newline='') as csvIn:
            wsI = csv.reader(csvIn,delimiter=',',quotechar='"')
            for row in wsI:
                a = row[0]
                if ( a == activity ):
                    thisID = row[4]
                    p = findAttendance(thisID,collated)
                    row[8] = p
                wsO.writerow(row)

# overwrite original file with temp file
    shutil.move(tFn, activityFn)
  

"""
return information about whether a particular student for a session was present/absent

input : student IDs (string), collated Dict
output : "Y" if student present, "N" if absent, "Not found" if id is not in dict
"""

def findAttendance(IDs,collated):
    
    ID = int(IDs)
    if ID in collated:
        v = collated[ID]
        if v["Present"]:
            return("Y")
        else:
            return("N")
    else:
        return("Not found")
        
    
"""
From the name of the TunrningPoint file and course
identify the acivitity and CRN from the relevant file

input : clickerFN, course, activityFN (csv file)
output: activity (string), CRN (string)
"""
def getActivityCRN(clickerFn,course,activityFn):
    AR = getActivityRoot(clickerFn)

    with open(activityFn, newline='') as csvfile:
        ws = csv.reader(csvfile,delimiter=',',quotechar='"')
        for row in ws:
            c = row[1]
            if ( c == course ):
                if row[0].find(AR) > -1:
                    return((row[0], row[3])) 
        return(None,None)        

   
"""
Update the register report for a single clicker file

"""    
def updateRegisterReportForAClickerFile(clickerFn,course, outputFolder, studentFn,activityFn):
    (activity,CRN) = getActivityCRN(clickerFn,course,activityFn)
#  Don't do anything if you cannot find the right activity date/time
    if activity:
        clickerIDs = ClickerRegister(clickerFn)
        studentData = buildStudentDict(studentFn)
        studentIDsAttending,malformedIDs = selectIDStudentData(studentData,clickerIDs)
        studentsAttending = collateStudentsInLecture(studentData,studentData.keys(),studentIDsAttending,malformedIDs)
        updateRegisterReport(studentsAttending,activity,activityFn)
   
    
"""
Pick out all .tpzx files in a directory and update register report for all of them 
"""
    
def updateReportForAFolder(clickerDir,course, outputFolder="/Users/upac004/Downloads/", studentFn="/Users/upac004/Downloads/Copy of Clicker Data CT Fixed (2).xlsx",activityFn="/Users/upac004/Downloads/AttendanceRegisterReport_Term2.csv"):
    
    with os.scandir(clickerDir) as it:
        for entry in it:
            if re.search(r'.tpzx',entry.name):
                print("Processing " + entry.name)
                updateRegisterReportForAClickerFile(entry.path,course, outputFolder, studentFn,activityFn)

                
                
sFn = "/Users/upac004/Downloads/Copy of Clicker Data CT Fixed (2).xlsx"
aFn = "/Users/upac004/Downloads/AttendanceRegisterReport_Term2.csv"
                
            

    
